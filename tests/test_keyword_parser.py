import io
import pytest
import openpyxl

from app.services.keyword_parser import parse_keyword_xlsx, parse_stop_word_xlsx


def _make_xlsx(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_single_col_xlsx(words: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for w in words:
        ws.append([w])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_basic_rows():
    data = _make_xlsx([
        ("ABM", '"ABM маркетинг", "Account based marketing"'),
        ("Конкуренты", '"конкурент", "аналог"'),
    ])
    result = parse_keyword_xlsx(data)
    assert result == [
        {"group": "ABM", "keywords": ["ABM маркетинг", "Account based marketing"]},
        {"group": "Конкуренты", "keywords": ["конкурент", "аналог"]},
    ]


def test_skips_empty_group_name():
    data = _make_xlsx([
        ("ABM", '"kw1"'),
        ("", '"kw2"'),
        (None, '"kw3"'),
    ])
    result = parse_keyword_xlsx(data)
    assert len(result) == 1
    assert result[0]["group"] == "ABM"


def test_empty_keywords_cell():
    data = _make_xlsx([
        ("ABM", ""),
    ])
    result = parse_keyword_xlsx(data)
    assert result == [{"group": "ABM", "keywords": []}]


def test_fewer_than_two_columns_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ABM"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError, match="at least 2 columns"):
        parse_keyword_xlsx(buf.getvalue())


def test_single_keyword_no_comma():
    data = _make_xlsx([
        ("Group", '"только один"'),
    ])
    result = parse_keyword_xlsx(data)
    assert result == [{"group": "Group", "keywords": ["только один"]}]


def test_parse_stop_words_basic():
    data = _make_single_col_xlsx(["конкурент", "тендер", "вакансия"])
    result = parse_stop_word_xlsx(data)
    assert result == ["конкурент", "тендер", "вакансия"]


def test_parse_stop_words_skips_empty():
    data = _make_single_col_xlsx(["слово", "", None, "другое"])
    result = parse_stop_word_xlsx(data)
    assert result == ["слово", "другое"]


def test_parse_stop_words_empty_file():
    data = _make_single_col_xlsx([])
    result = parse_stop_word_xlsx(data)
    assert result == []


def test_parse_stop_words_strips_whitespace():
    data = _make_single_col_xlsx(["  пробел  ", "чистый"])
    result = parse_stop_word_xlsx(data)
    assert result == ["пробел", "чистый"]
