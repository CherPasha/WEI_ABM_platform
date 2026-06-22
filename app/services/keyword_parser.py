import io

import openpyxl


def parse_keyword_xlsx(file_bytes: bytes) -> list[dict]:
    """
    Parse an Excel file where each row is one keyword:
      col 0 — keyword group name (may repeat across rows)
      col 1 — single keyword

    Returns list of {"group": str, "keywords": list[str]}, one entry per group.
    Raises ValueError if file has fewer than 2 columns.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    if ws.max_column is not None and ws.max_column < 2:
        raise ValueError("File must have at least 2 columns")

    groups: dict[str, list[str]] = {}
    group_order: list[str] = []

    for row in ws.iter_rows(values_only=True):
        if len(row) < 2:
            continue
        raw_group = row[0]
        raw_kw = row[1]

        group_name = str(raw_group).strip() if raw_group is not None else ""
        keyword = str(raw_kw).strip() if raw_kw is not None else ""

        if not group_name or not keyword:
            continue

        if group_name not in groups:
            groups[group_name] = []
            group_order.append(group_name)

        for token in keyword.split(","):
            kw = token.strip().strip('"')
            if kw:
                groups[group_name].append(kw)

    wb.close()
    return [{"group": g, "keywords": groups[g]} for g in group_order]


def parse_stop_word_xlsx(file_bytes: bytes) -> list[str]:
    """
    Parse an Excel file with one column of stop words, one word per row.

    Returns list of non-empty stripped strings.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    results = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        raw_word = row[0]
        word = str(raw_word).strip() if raw_word is not None else ""
        if word:
            results.append(word)
    wb.close()
    return results


def parse_roles_xlsx(file_bytes: bytes) -> list[str]:
    """
    Parse an Excel file with one column of role names, one role per row.

    Returns list of non-empty stripped strings.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    results = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        raw_role = row[0]
        role = str(raw_role).strip() if raw_role is not None else ""
        if role:
            results.append(role)
    wb.close()
    return results
